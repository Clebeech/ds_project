from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
import re


BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "processed"

EXCEL_DATE_ORIGIN = "1899-12-30"


def ensure_output_dir() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def load_excel(path: Path, sheet: str | int = 0) -> pd.DataFrame:
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        if isinstance(sheet, int):
            ws = wb.worksheets[sheet]
        else:
            ws = wb[sheet]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    if not rows:
        return pd.DataFrame()
    header = []
    for idx, value in enumerate(rows[0]):
        header.append(str(value) if value is not None else f"column_{idx}")
    data = rows[1:]
    df = pd.DataFrame(data, columns=header)
    df = df.apply(lambda col: col.map(lambda x: x.replace("\xa0", " ").strip() if isinstance(x, str) else x))
    return df


def strip_to_float(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("%", "", regex=False)
        .str.replace("，", "", regex=False)
        .str.strip()
    )
    cleaned = cleaned.replace({"": pd.NA, "nan": pd.NA, "None": pd.NA})
    return pd.to_numeric(cleaned, errors="coerce").astype("float64")


def convert_numeric_columns(df: pd.DataFrame, skip: Iterable[str] = ()) -> pd.DataFrame:
    numeric_pattern = re.compile(r"^[-+]?\d+(\.\d+)?$")
    skip_set = {col.lower() for col in skip}
    for col in df.columns:
        if col.lower() in skip_set:
            continue
        sample = df[col].dropna().astype(str).str.replace(",", "", regex=False).str.strip()
        if sample.empty:
            continue
        numeric_ratio = sample.apply(lambda v: bool(numeric_pattern.fullmatch(v))).mean()
        if numeric_ratio >= 0.7:
            df[col] = strip_to_float(df[col])
    return df


def convert_excel_date(series: pd.Series) -> pd.Series:
    def _convert(value: str | float | int) -> pd.Timestamp | pd.NA:
        if value is None:
            return pd.NA
        if isinstance(value, str):
            text = value.strip()
            if text == "":
                return pd.NA
            try:
                numeric = float(text)
            except ValueError:
                parsed = pd.to_datetime(text, errors="coerce")
                return parsed if not pd.isna(parsed) else pd.NA
        elif pd.isna(value):
            return pd.NA
        else:
            numeric = value
        try:
            numeric = float(numeric)
        except (TypeError, ValueError):
            try:
                parsed = pd.to_datetime(numeric, errors="coerce")
            except Exception:
                return pd.NA
            return parsed if not pd.isna(parsed) else pd.NA
        return pd.to_datetime(numeric, unit="d", origin=EXCEL_DATE_ORIGIN, errors="coerce")

    converted = series.map(_convert)
    return converted


def normalize_numeric_code(series: pd.Series, width: int | None = None) -> pd.Series:
    def _format(value: object) -> pd.NA | str:
        if value is None or pd.isna(value):
            return pd.NA
        text = str(value).strip()
        if not text:
            return pd.NA
        # 去除全角空格与尾随 .0
        text = text.replace("\u3000", "").replace(".0", "")
        digits = re.sub(r"[^\d]", "", text)
        if not digits:
            return pd.NA
        if width is not None and len(digits) <= width:
            digits = digits.zfill(width)
        return digits

    return series.map(_format).astype("string")


def clean_county_core() -> None:
    path = DATA_DIR / "各县基本数据" / "综合数据.xlsx"
    df = load_excel(path)
    if "CountyCode" in df.columns:
        df["CountyCode"] = normalize_numeric_code(df["CountyCode"], width=6)
    if "county_code" in df.columns:
        df["county_code"] = normalize_numeric_code(df["county_code"], width=6)
    df = convert_numeric_columns(
        df,
        skip=[
            "CountyCode",
            "地区名称",
            "所属省份",
            "所属城市",
        ],
    )
    for col in ["经度", "纬度"]:
        df[col] = strip_to_float(df[col])
    df.rename(columns={"县级行政区划代码": "CountyCode"}, inplace=True)
    df.to_csv(OUTPUT_DIR / "county_profile.csv", index=False)


def clean_basic_tables() -> None:
    mapping = {
        "人口.xlsx": "county_population.csv",
        "地区生产总值及指数.xlsx": "county_gdp.csv",
        "城乡居民收入.xlsx": "county_income.csv",
        "交通运输与邮电.xlsx": "county_transport_post.csv",
        "财政收支.xlsx": "county_finance_budget_raw.csv",
        "金融机构存贷款.xlsx": "county_financial_services.csv",
    }
    for file_name, output_name in mapping.items():
        path = DATA_DIR / "各县基本数据" / file_name
        df = load_excel(path)
        if "CountyCode" in df.columns:
            df["CountyCode"] = normalize_numeric_code(df["CountyCode"], width=6)
        df = convert_numeric_columns(
            df,
            skip=["CountyName", "PrefLevCity", "Province", "Project", "AreaID", "CountyCode"],
        )
        df.to_csv(OUTPUT_DIR / output_name, index=False)


def clean_archive_tables() -> None:
    archive_dir = DATA_DIR / "archive"

    # 农产品产量
    agri_output = load_excel(archive_dir / "中国各县域农产品产量数据.xlsx")
    for col in ["县域代码"]:
        if col in agri_output.columns:
            agri_output[col] = normalize_numeric_code(agri_output[col], width=6)
    agri_output = convert_numeric_columns(
        agri_output,
        skip=["县域名称", "所属地级市", "所属省份", "产品种类或名称", "单位", "地区编码ID", "县域代码"],
    )
    agri_output.to_csv(OUTPUT_DIR / "county_agri_output.csv", index=False)

    # 农作物播种面积
    crop_area = load_excel(archive_dir / "中国各县域农作物播种面积数据.xlsx")
    for col in ["县域代码"]:
        if col in crop_area.columns:
            crop_area[col] = normalize_numeric_code(crop_area[col], width=6)
    crop_area = convert_numeric_columns(
        crop_area,
        skip=["县域名称", "所属地级市", "所属省份", "农作物种类或名称", "地区编码ID", "县域代码"],
    )
    crop_area.to_csv(OUTPUT_DIR / "county_crop_area.csv", index=False)

    # 地形起伏度
    relief = load_excel(archive_dir / "中国各县域地形起伏度数据.xlsx", sheet="区县")
    relief = convert_numeric_columns(
        relief,
        skip=["县域名称", "所属省份", "地区等级"],
    )
    relief.to_csv(OUTPUT_DIR / "county_relief.csv", index=False)

    # 贫困县名单
    poverty = load_excel(archive_dir / "全国832个国家级贫困县名单.xlsx", sheet="整理数据")
    poverty = convert_numeric_columns(
        poverty,
        skip=["县域名称", "所属地域", "所属省份", "所属城市"],
    )
    poverty.to_csv(OUTPUT_DIR / "poverty_county_list.csv", index=False)


def clean_mapping_tables() -> None:
    mapping_df = load_excel(DATA_DIR / "对应县.xlsx")
    mapping_df.rename(
        columns={
            "负责县域ID": "county_code",
            "对应县": "county_alias",
            "地区": "region_detail",
        },
        inplace=True,
    )
    if "county_code" in mapping_df.columns:
        mapping_df["county_code"] = normalize_numeric_code(mapping_df["county_code"], width=6)
    mapping_df.to_csv(OUTPUT_DIR / "county_case_mapping.csv", index=False)


def clean_surveyor_tables() -> None:
    surveyor_df = load_excel(DATA_DIR / "调研员表.xlsx")
    if "调研员ID" in surveyor_df.columns:
        surveyor_df["调研员ID"] = surveyor_df["调研员ID"].astype(str).str.strip()
    if "所属分队ID" in surveyor_df.columns:
        surveyor_df["所属分队ID"] = surveyor_df["所属分队ID"].astype(str).str.strip()
    if "负责县域ID" in surveyor_df.columns:
        surveyor_df["负责县域ID"] = normalize_numeric_code(surveyor_df["负责县域ID"], width=6)
    surveyor_df = convert_numeric_columns(
        surveyor_df,
        skip=[
            "调研员ID",
            "姓名",
            "性别",
            "所属院系",
            "学历",
            "专业方向",
            "联系电话",
            "电子邮箱",
            "所属分队ID",
            "负责县域ID",
            "调研角色",
            "参与调研批次",
            "调研专长",
            "培训完成状态",
            "设备领用状态",
            "备注",
        ],
    )
    for col in ["调研开始时间", "调研结束时间"]:
        if col in surveyor_df.columns:
            surveyor_df[col] = convert_excel_date(surveyor_df[col])
    surveyor_df.to_csv(OUTPUT_DIR / "surveyors.csv", index=False)

    interviews_df = load_excel(DATA_DIR / "访谈记录.xlsx")
    if "县id" in interviews_df.columns:
        interviews_df["县id"] = normalize_numeric_code(interviews_df["县id"], width=6)
    if "调研人id" in interviews_df.columns:
        interviews_df["调研人id"] = interviews_df["调研人id"].astype(str).str.strip()
    interviews_df = convert_numeric_columns(
        interviews_df,
        skip=[
            "访谈记录id",
            "调研人id",
            "县id",
            "访谈对象id",
            "受访人姓名",
            "访谈内容",
            "访谈地点",
        ],
    )
    interviews_df["访谈时间"] = convert_excel_date(interviews_df["访谈时间"])
    interviews_df["受访人信息"] = interviews_df["受访人信息"].fillna("").str.strip()
    interviews_df["访谈内容"] = interviews_df["访谈内容"].fillna("").str.strip()
    interviews_df.to_csv(OUTPUT_DIR / "interviews.csv", index=False)


def main() -> None:
    start = datetime.now()
    ensure_output_dir()
    clean_county_core()
    clean_basic_tables()
    clean_archive_tables()
    clean_mapping_tables()
    clean_surveyor_tables()
    elapsed = datetime.now() - start
    print(f"Data cleaning completed in {elapsed.total_seconds():.2f}s. Output -> {OUTPUT_DIR}")


if __name__ == "__main__":
    main()


